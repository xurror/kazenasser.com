
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook

from article import Article


def export_news(news_per_source: List[Dict[str, List[Article]]]):
    print("Exporting news to Excel...")
    
    filename = f'{datetime.now().strftime("%d.%m.%Y")}-news.xlsx'

    wb = Workbook()
    ws = wb.active
    
    ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1

    if ws is not None:
        ws.title = 'Competitor News'
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 85
        ws.column_dimensions['D'].width = 120
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 50

        row = 1
        for source in news_per_source:

            ws.append([source])
            # make the whole row bold
            for cell in ws[f'A{row}:F{row}'][0]:
                cell.font = cell.font.copy(bold=True)
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            ws.append(['No', 'Date', 'Title', 'Description', 'Domain', 'Link'])
            # make the whole row bold
            for cell in ws[f'A{row}:F{row}'][0]:
                cell.font = cell.font.copy(bold=True)
            row += 1

            news: List[Article] = Article.sort_by_date(news_per_source[source])
            
            count = 1
            for article in news:
                ws.append([
                    count,
                    article.get_date_str(),
                    article.title,
                    article.description,
                    article.domain,
                    article.link
                ])

                row += 1
                count += 1
            ws.append([])
            row += 1

        wb.save(filename=filename)

    print("Exporting news to Excel... Done!")
    return filename
